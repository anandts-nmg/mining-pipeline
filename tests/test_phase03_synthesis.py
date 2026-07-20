"""Phase 03 (Geological, Metallogenic & CMCS Synthesis) tests against synthetic fixtures.

Phase 03 consumes Phase 01's licence-boundary AOI (for the CMCS buffer), so the real-run
test runs Phase 00 (working copies) and Phase 01 (boundary + master GPKG) first. The #68
mineralized-point XLSX ingest is expected to SKIP gracefully on the synthetic placeholder
(the raw archive writes a text placeholder, not a real workbook).
"""

from __future__ import annotations

import openpyxl

from buduunkhad.core import paths, vector_io
from buduunkhad.core.gates import GateStatus
from buduunkhad.phases.base import RunContext
from buduunkhad.phases.phase00_archive import Phase00Archive
from buduunkhad.phases.phase01_data_audit import Phase01DataAudit
from buduunkhad.phases.phase03_geology_synthesis import (
    EVIDENCE_FIELDS,
    EVIDENCE_LAYERS,
    Phase03GeologySynthesis,
)

_MANDATORY = set(EVIDENCE_FIELDS)  # 13 provenance fields + feature_id = 14 columns


def _ctx(config, register, *, dry_run=False):
    return RunContext(config=config, register=register, run_id="test03", dry_run=dry_run)


def _run_real(config, register):
    ctx = _ctx(config, register)
    Phase00Archive().run(ctx)
    p1 = Phase01DataAudit()
    p1.prepare(ctx)
    p1.run(ctx)
    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    result = phase.run(ctx)
    return ctx, phase, result


def _evidence_gpkg(config):
    pdir = paths.phase_dir(config.output_root, "03")
    gpkgs = list((pdir / "09_Geological_Evidence_Layers_GPKG").glob("*Geological_Evidence*.gpkg"))
    assert gpkgs, "evidence GPKG not created"
    return gpkgs[0]


def test_phase03_scaffolds_twelve_folders(project):
    config, register, _tmp = project
    ctx = _ctx(config, register, dry_run=True)
    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    pdir = paths.phase_dir(config.output_root, "03")
    subs = phase.custom_subfolders
    assert subs is not None
    for sub in subs:
        assert (pdir / sub).is_dir(), f"missing folder {sub}"
    assert len(subs) == 12


def test_phase03_evidence_schema_fields_and_geometry(project):
    config, register, _tmp = project
    ctx = _ctx(config, register, dry_run=True)
    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    phase.run(ctx)

    gpkg = _evidence_gpkg(config)
    layers = vector_io.list_gpkg_layers(gpkg)
    expected = {name for name, _g, _p in EVIDENCE_LAYERS}
    assert set(layers) == expected
    assert len(layers) == 17

    import fiona

    for name, geom, _prefix in EVIDENCE_LAYERS:
        assert geom.startswith("Multi")  # layers are Multi* so promoted ingests append cleanly
        with fiona.open(gpkg, layer=name) as src:
            schema = src.schema
            assert schema is not None
            props = set(schema["properties"])
            assert props == _MANDATORY, f"{name} field mismatch: {props ^ _MANDATORY}"
            assert "feature_id" in props
            assert schema["geometry"] == geom, f"{name} geometry mismatch"


def test_phase03_feature_id_generation():
    fid = Phase03GeologySynthesis.feature_id
    assert fid("BUD-MIN", 1) == "BUD-MIN-0001"
    assert fid("BUD-GEO50", 42) == "BUD-GEO50-0042"
    assert fid("BUD-STR", 12345) == "BUD-STR-12345"


def test_phase03_dry_run_templates_only(project):
    config, register, _tmp = project
    ctx = _ctx(config, register, dry_run=True)
    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    result = phase.run(ctx)
    assert result.status == "dry-run"

    pdir = paths.phase_dir(config.output_root, "03")
    # deposit model template + score matrix + registers present
    assert list(
        (pdir / "10_Preliminary_Deposit_Model_03A").glob("*Preliminary_Deposit_Model*.docx")
    )
    assert list((pdir / "10_Preliminary_Deposit_Model_03A").glob("*score_matrix*.xlsx"))
    assert list(
        (pdir / "02_Tectonic_Terrane_Context").glob("*Tectonic_Terrane_Context_Register*.xlsx")
    )
    assert list((pdir / "12_Phase3_QAQC_and_Handover").glob("*Phase3_QAQC_Log*.xlsx"))
    # schema present but empty (no ingested points, no CMCS buffer built)
    gpkg = _evidence_gpkg(config)
    assert len(vector_io.list_gpkg_layers(gpkg)) == 17
    # no CMCS buffer gpkg in dry-run
    assert not list(
        (pdir / "08_CMCS_MRPAM_Buffer_Check_5km_10km_20km").glob("*CMCS_MRPAM_Buffer*.gpkg")
    )


def test_phase03_score_matrix_shape(project):
    config, register, _tmp = project
    ctx = _ctx(config, register, dry_run=True)
    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    phase.run(ctx)
    pdir = paths.phase_dir(config.output_root, "03")
    matrix = next((pdir / "10_Preliminary_Deposit_Model_03A").glob("*score_matrix*.xlsx"))
    wb = openpyxl.load_workbook(matrix)
    ws = wb.active
    assert ws is not None
    header = [c.value for c in ws[1]]
    # criterion + max_points + 6 models = 8 columns
    assert len(header) == 8
    # 8 criteria + total row = 9 data rows
    assert ws.max_row == 1 + 9


def test_phase03_cmcs_buffer_four_rings(raw_archive):
    config, register, _raw = raw_archive
    ctx, phase, result = _run_real(config, register)
    assert result.status == "ok"
    assert phase._cmcs_rings == 4  # v8/v9 Step 7 adds the 25 km ring

    pdir = paths.phase_dir(config.output_root, "03")
    buffers = list(
        (pdir / "08_CMCS_MRPAM_Buffer_Check_5km_10km_20km_25km").glob("*CMCS_MRPAM_Buffer*.gpkg")
    )
    assert buffers, "CMCS buffer gpkg not written"
    gdf = vector_io.read_layer(buffers[0], "cmcs_mrpam_buffer")
    assert len(gdf) == 4
    assert sorted(gdf["distance_m"]) == [5000, 10000, 20000, 25000]
    assert (gdf["validation_status"] == "Historical only").all()


def test_phase03_step7a_standalone_25km_buffer(raw_archive):
    # v8/v9 Step 7A: a standalone 25 km coverage buffer + the coverage-check register are emitted
    # even when #68 is the synthetic placeholder (buffer is boundary-derived; register is a template).
    config, register, _raw = raw_archive
    ctx, phase, result = _run_real(config, register)
    assert result.status == "ok"
    cmcs = (
        paths.phase_dir(config.output_root, "03") / "08_CMCS_MRPAM_Buffer_Check_5km_10km_20km_25km"
    )
    b25 = list(cmcs.glob("*Buffer_25km*.gpkg"))
    assert b25, "standalone 25 km coverage buffer not written"
    g = vector_io.read_layer(b25[0], "license_boundary_buffer_25km")
    assert len(g) == 1
    assert list(cmcs.glob("*25km_Near_Occurrence_Coverage_Check_Register*.xlsx")), (
        "25 km coverage-check register not emitted"
    )


def test_phase03_step7a_coverage_from_68(raw_archive):
    # A real #68 workbook (points inside the AOI) exercises the coverage check: per-occurrence
    # within-20/25 km flags in the register + the within-25 km selection layer written.
    import openpyxl as pyxl

    config, register, _raw = raw_archive
    ctx = _ctx(config, register)
    Phase00Archive().run(ctx)
    p1 = Phase01DataAudit()
    p1.prepare(ctx)
    p1.run(ctx)

    rec = ctx.record_by_no(68)
    wc = paths.phase_dir(config.output_root, "00") / rec.evidence_group / rec.filename
    wc.parent.mkdir(parents=True, exist_ok=True)
    wb = pyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "commodity", "lon", "lat"])
    ws.append(["OccA", "Au", 96.50, 45.50])
    ws.append(["OccB", "Cu", 96.55, 45.52])
    wb.save(wc)

    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    phase.run(ctx)

    cmcs = (
        paths.phase_dir(config.output_root, "03") / "08_CMCS_MRPAM_Buffer_Check_5km_10km_20km_25km"
    )
    sel = list(cmcs.glob("*near_mineral_occurrences_within_25km*.gpkg"))
    assert sel, "within-25 km selection layer not written"
    g = vector_io.read_layer(sel[0], "near_mineral_occurrences_within_25km")
    assert len(g) == 2

    reg = list(cmcs.glob("*25km_Near_Occurrence_Coverage_Check_Register*.xlsx"))
    assert reg
    ws2 = openpyxl.load_workbook(reg[0]).active
    assert ws2 is not None
    rows = list(ws2.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]
    assert len(data) == 2, "coverage register should carry both ingested occurrences"
    assert "distance_to_boundary_km" in header
    assert all(r[header.index("within_25km")] == "yes" for r in data)


def test_phase03_xlsx_ingest_skips_gracefully(raw_archive):
    """The synthetic #68 raw is a text placeholder, not a real workbook: ingest must skip,
    record a note, and not crash — leaving 0 mineralized points."""
    config, register, _raw = raw_archive
    ctx, phase, result = _run_real(config, register)
    assert result.status == "ok"
    assert phase._mineralized_points == 0
    assert any("#68" in n for n in phase._notes)
    # the mineralized layer still exists in the evidence GPKG (empty)
    gpkg = _evidence_gpkg(config)
    assert "mineralized_points_point" in vector_io.list_gpkg_layers(gpkg)


def test_phase03_xlsx_ingest_real_workbook(raw_archive):
    """A real #68 workbook with lon/lat columns ingests, reprojects and generates BUD- IDs."""
    import openpyxl as pyxl

    config, register, _raw = raw_archive
    ctx = _ctx(config, register)
    Phase00Archive().run(ctx)
    p1 = Phase01DataAudit()
    p1.prepare(ctx)
    p1.run(ctx)

    # overwrite the #68 working copy with a real workbook (lon/lat over the AOI)
    rec = ctx.record_by_no(68)
    wc = paths.phase_dir(config.output_root, "00") / rec.evidence_group / rec.filename
    wc.parent.mkdir(parents=True, exist_ok=True)
    wb = pyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "commodity", "lon", "lat"])
    ws.append(["OccA", "Au", 96.50, 45.50])
    ws.append(["OccB", "Cu", 96.55, 45.52])
    wb.save(wc)

    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    phase.run(ctx)
    assert phase._mineralized_points == 2

    gpkg = _evidence_gpkg(config)
    gdf = vector_io.read_layer(gpkg, "mineralized_points_point")
    assert len(gdf) == 2
    assert gdf.crs is not None and gdf.crs.to_epsg() == config.target_epsg
    assert sorted(gdf["feature_id"]) == ["BUD-MIN-0001", "BUD-MIN-0002"]
    assert (gdf["validation_status"] == "Historical only").all()


def test_phase03_xlsx_ingest_mongolian_dms(raw_archive):
    """The real #68 register uses Mongolian headers (Уртраг/Өргөрөг) with DMS coordinate
    strings (96°41'16"). Those must be detected, parsed to decimal degrees, and reprojected."""
    import openpyxl as pyxl

    config, register, _raw = raw_archive
    ctx = _ctx(config, register)
    Phase00Archive().run(ctx)
    p1 = Phase01DataAudit()
    p1.prepare(ctx)
    p1.run(ctx)

    rec = ctx.record_by_no(68)
    wc = paths.phase_dir(config.output_root, "00") / rec.evidence_group / rec.filename
    wc.parent.mkdir(parents=True, exist_ok=True)
    wb = pyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["№", "Цэгийн дугаар", "Уртраг", "Өргөрөг"])
    ws.append([1, "MO-1", "96°41'16\"", "45°57'32\""])
    ws.append([2, "MO-2", "96°40'46\"", "45°59'27\""])
    wb.save(wc)

    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    phase.run(ctx)
    assert phase._mineralized_points == 2

    gpkg = _evidence_gpkg(config)
    gdf = vector_io.read_layer(gpkg, "mineralized_points_point")
    assert len(gdf) == 2
    assert gdf.crs is not None and gdf.crs.to_epsg() == config.target_epsg
    assert sorted(gdf["feature_id"]) == ["BUD-MIN-0001", "BUD-MIN-0002"]
    # DMS parsed to lon ~96.7 / lat ~46.0 then reprojected -> finite metre-scale UTM
    coords = gdf.geometry.get_coordinates()
    assert (coords["x"].abs() < 1e7).all() and (coords["y"].abs() < 1e7).all()


def test_phase03_occurrence_register_populated_from_68(raw_archive):
    """#68 source attributes are preserved into the occurrence registers (not discarded): the
    Mineral_Occurrences_Register carries name/commodity/coords, and unmapped columns (e.g. rock)
    survive in the cross-reference note."""
    import openpyxl as pyxl

    config, register, _raw = raw_archive
    ctx = _ctx(config, register)
    Phase00Archive().run(ctx)
    p1 = Phase01DataAudit()
    p1.prepare(ctx)
    p1.run(ctx)

    rec = ctx.record_by_no(68)
    wc = paths.phase_dir(config.output_root, "00") / rec.evidence_group / rec.filename
    wc.parent.mkdir(parents=True, exist_ok=True)
    wb = pyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "commodity", "lon", "lat", "rock"])
    ws.append(["OccA", "Au", 96.50, 45.50, "granite"])
    ws.append(["OccB", "Cu", 96.55, 45.52, "diorite"])
    wb.save(wc)

    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    phase.run(ctx)

    pdir = paths.phase_dir(config.output_root, "03")
    reg_files = list(
        (pdir / "05_Local_Geology_Occurrence_1M50K").glob("*Mineral_Occurrences_Register*.xlsx")
    )
    assert reg_files, "occurrence register not emitted"
    ws2 = openpyxl.load_workbook(reg_files[0]).active
    assert ws2 is not None
    rows = list(ws2.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]
    assert len(data) == 2, "occurrence register should carry the 2 ingested #68 points"
    by_name = {r[header.index("occurrence_name")]: r for r in data}
    assert set(by_name) == {"OccA", "OccB"}
    assert by_name["OccA"][header.index("commodity")] == "Au"
    assert int(str(by_name["OccA"][header.index("source_raw_input_no")])) == 68
    assert by_name["OccA"][header.index("validation_status")] == "Historical only"
    # eastings are metre-scale UTM, not lon/lat degrees
    assert float(str(by_name["OccA"][header.index("easting_32647")])) > 1000

    # the unmapped 'rock' column is preserved verbatim in the cross-reference note
    xref_files = list(
        (pdir / "07_Occurrence_Register_and_Coordinate_QAQC").glob(
            "*Occurrence_CrossReference*.xlsx"
        )
    )
    assert xref_files
    wsx = openpyxl.load_workbook(xref_files[0]).active
    assert wsx is not None
    xrows = list(wsx.iter_rows(values_only=True))
    xheader, xdata = xrows[0], xrows[1:]
    notes = " ".join(str(r[xheader.index("note")]) for r in xdata)
    assert "rock=granite" in notes and "rock=diorite" in notes
    assert all(r[xheader.index("in_68_xlsx")] == "yes" for r in xdata)


def test_phase03_method_note_published_location(raw_archive):
    """The Phase 3 method note is written into the handover folder (not the publish-excluded
    01_Input_Working_Copy) so it ships in the deliverable package."""
    from buduunkhad.core.publish import collect_deliverables

    config, register, _raw = raw_archive
    _ctx_, _phase, _result = _run_real(config, register)
    pdir = paths.phase_dir(config.output_root, "03")
    note_name = f"{config.register_prefix}_Phase3_Method_Note.md"
    note = pdir / "12_Phase3_QAQC_and_Handover" / note_name
    assert note.exists()
    note_text = note.read_text(encoding="utf-8")
    assert "preliminary 17-layer support-evidence schema/package" in note_text
    assert "authoritative 17-layer" not in note_text
    assert "Layer presence is not evidence completeness" in note_text
    assert not (pdir / "01_Input_Working_Copy" / note_name).exists()
    assert note_name in {p.name for p in collect_deliverables(config.output_root)}


def test_phase03_ingest_human_layer(raw_archive):
    """A human-digitized layer dropped into a phase folder is ingested into the evidence GPKG."""
    import geopandas as gpd
    from shapely.geometry import LineString

    config, register, _raw = raw_archive
    ctx = _ctx(config, register)
    Phase00Archive().run(ctx)
    p1 = Phase01DataAudit()
    p1.prepare(ctx)
    p1.run(ctx)

    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)

    pdir = paths.phase_dir(config.output_root, "03")
    human = pdir / "04_Regional_Geology_Mineral_1M200K" / "human_faults.gpkg"
    gdf = gpd.GeoDataFrame(
        {"note": ["fault A"]},
        geometry=[LineString([(400000, 5000000), (401000, 5001000)])],
        crs=f"EPSG:{config.target_epsg}",
    )
    vector_io.write_layer(gdf, human, layer="faults_structures_line")

    phase.run(ctx)
    assert "faults_structures_line" in phase._ingested_layers
    gpkg = _evidence_gpkg(config)
    ingested = vector_io.read_layer(gpkg, "faults_structures_line")
    assert len(ingested) == 1
    assert ingested["feature_id"].iloc[0] == "BUD-STR-0001"
    assert ingested["validation_status"].iloc[0] == "Historical only"


def test_phase03_keeps_ai_handoff_evidence_out_of_legacy_schema_normalization(raw_archive):
    """Rich AI review lineage must not be stripped by the legacy 14-column normalizer."""

    import geopandas as gpd
    from shapely.geometry import LineString, MultiLineString

    config, register, _raw = raw_archive
    ctx = _ctx(config, register)
    Phase00Archive().run(ctx)
    phase1 = Phase01DataAudit()
    phase1.prepare(ctx)
    phase1.run(ctx)
    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    source = paths.phase_dir(config.output_root, "03") / "01_Input_Working_Copy"
    source.mkdir(parents=True, exist_ok=True)
    tagged = source / "accepted-ai-evidence.gpkg"
    gpd.GeoDataFrame(
        {
            "feature_id": ["BUD-STR-ACCEPTED"],
            "proposal_state": ["AI_DRAFT"],
            "review_state": ["HUMAN_REVIEWED"],
            "evidence_state": ["ACCEPTED_EVIDENCE"],
            "request_fingerprint": ["a" * 64],
            "geometry": [MultiLineString([LineString([(500000, 5200000), (500100, 5200100)])])],
        },
        crs=f"EPSG:{config.target_epsg}",
    ).to_file(tagged, layer="faults_structures_line", driver="GPKG")

    phase.run(ctx)
    evidence = vector_io.read_layer(_evidence_gpkg(config), "faults_structures_line")
    assert len(evidence) == 0
    assert any("kept separate from legacy normalization" in note for note in phase._notes)


def test_phase03_qaqc_items_and_gate_go(raw_archive):
    config, register, _raw = raw_archive
    ctx, phase, _result = _run_real(config, register)
    report = phase.qaqc(ctx)
    items = [i.item for i in report.items]
    assert any("Master GIS" in i for i in items)
    assert any("coordinate QA/QC" in i for i in items)
    assert any("CMCS/MRPAM 5/10/20/25 km buffer" in i for i in items)
    assert any("Preliminary Deposit Model" in i for i in items)
    assert any("Historical only" in i for i in items)
    assert any("preliminary support-evidence schema/package emitted" in i for i in items)
    assert not any("ready for Phase 4" in i for i in items)
    schema_item = next(
        item for item in report.items if "preliminary support-evidence schema/package" in item.item
    )
    assert "not evidence completeness or scientific approval" in schema_item.note
    assert not report.has_failures

    decision = phase.gate(report, ctx)
    assert decision.status is GateStatus.GO, decision.reason


def test_phase03_human_layer_reprojected_from_4326(raw_archive):
    """A human layer supplied in EPSG:4326 is reprojected to the target CRS on ingest (invariant #4)."""
    import geopandas as gpd
    from shapely.geometry import LineString

    config, register, _raw = raw_archive
    ctx = _ctx(config, register)
    Phase00Archive().run(ctx)
    p1 = Phase01DataAudit()
    p1.prepare(ctx)
    p1.run(ctx)

    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    pdir = paths.phase_dir(config.output_root, "03")
    human = pdir / "05_Local_Geology_Occurrence_1M50K" / "human_faults_4326.gpkg"
    gdf = gpd.GeoDataFrame(
        {"note": ["fault B"]},
        geometry=[LineString([(96.50, 45.50), (96.52, 45.52)])],
        crs="EPSG:4326",
    )
    vector_io.write_layer(gdf, human, layer="faults_structures_line")

    phase.run(ctx)
    gpkg = _evidence_gpkg(config)
    ingested = vector_io.read_layer(gpkg, "faults_structures_line")
    assert len(ingested) == 1
    assert ingested.crs is not None and ingested.crs.to_epsg() == config.target_epsg
    # coordinates are now metre-scale UTM eastings, not lon/lat degrees
    coords = ingested.geometry.get_coordinates()
    assert not coords.empty and (coords["x"].abs() > 1000).all()


def test_phase03_xlsx_ingest_utm_easting_northing(raw_archive):
    """A #68 workbook in UTM easting/northing (metre-scale) is NOT mislabeled as lon/lat -> no (inf, inf)."""
    import openpyxl as pyxl

    config, register, _raw = raw_archive
    ctx = _ctx(config, register)
    Phase00Archive().run(ctx)
    p1 = Phase01DataAudit()
    p1.prepare(ctx)
    p1.run(ctx)

    rec = ctx.record_by_no(68)
    wc = paths.phase_dir(config.output_root, "00") / rec.evidence_group / rec.filename
    wc.parent.mkdir(parents=True, exist_ok=True)
    wb = pyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "commodity", "easting", "northing"])
    ws.append(["OccU", "Au", 305000.0, 5040000.0])
    ws.append(["OccV", "Cu", 310000.0, 5042000.0])
    wb.save(wc)

    phase = Phase03GeologySynthesis()
    phase.prepare(ctx)
    phase.run(ctx)
    assert phase._mineralized_points == 2

    gpkg = _evidence_gpkg(config)
    gdf = vector_io.read_layer(gpkg, "mineralized_points_point")
    assert len(gdf) == 2
    assert gdf.crs is not None and gdf.crs.to_epsg() == config.target_epsg
    coords = gdf.geometry.get_coordinates()
    assert (coords["x"].abs() < 1e7).all() and (
        coords["y"].abs() < 1e7
    ).all()  # finite, not (inf, inf)


def test_phase03_license_boundary_populated(raw_archive):
    """The evidence GPKG's license_boundary layer is populated from the Phase 01 boundary."""
    config, register, _raw = raw_archive
    _ctx_, phase, _result = _run_real(config, register)
    gpkg = _evidence_gpkg(config)
    boundary = vector_io.read_layer(gpkg, "license_boundary")
    assert len(boundary) >= 1
    assert boundary.crs is not None and boundary.crs.to_epsg() == config.target_epsg


def test_phase03_input_numbers():
    phase = Phase03GeologySynthesis()
    assert phase.input_numbers == [*range(1, 9), *range(53, 73)]
    assert 8 in phase.input_numbers
