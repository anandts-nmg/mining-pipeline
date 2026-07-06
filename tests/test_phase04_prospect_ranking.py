"""Phase 04 (Preliminary Prospect Delineation & Ranking) tests against synthetic fixtures.

Phase 04 consumes Phase 03's evidence GPKG (for the scoring grid), so the real-run helper chains
Phase 00 → 01 → 03 → 04. On the bare synthetic archive most evidence layers are empty, so the grid
scores ~0 and no candidate reaches the threshold — that path is exercised too. The delineation test
writes a real #68 workbook (so Phase 03 ingests points) and injects a geology contact + fault near
those points so cells there score (geology-contact 20 + structure 15 + geochem 15) and prospects form.
"""

from __future__ import annotations

import openpyxl

from buduunkhad.core import paths, vector_io
from buduunkhad.core.gates import GateStatus
from buduunkhad.phases.base import RunContext
from buduunkhad.phases.phase00_archive import Phase00Archive
from buduunkhad.phases.phase01_data_audit import Phase01DataAudit
from buduunkhad.phases.phase03_geology_synthesis import EVIDENCE_FIELDS, Phase03GeologySynthesis
from buduunkhad.phases.phase04_prospect_ranking import (
    PROSPECT_CRITERIA,
    Phase04ProspectRanking,
    classify,
)


def _ctx(config, register, *, dry_run=False):
    return RunContext(config=config, register=register, run_id="test04", dry_run=dry_run)


def _p04_dir(config):
    return paths.phase_dir(config.output_root, "04")


def _run_to_03(ctx):
    Phase00Archive().run(ctx)
    p1 = Phase01DataAudit()
    p1.prepare(ctx)
    p1.run(ctx)


def _evidence_gpkg_03(config):
    p03 = paths.phase_dir(config.output_root, "03") / "09_Geological_Evidence_Layers_GPKG"
    return next(p03.glob("*Geological_Evidence*.gpkg"))


def _write_68(ctx, config):
    """Overwrite the #68 working copy with a real workbook of 2 points inside the AOI."""
    import openpyxl as pyxl

    rec = ctx.record_by_no(68)
    wc = paths.phase_dir(config.output_root, "00") / rec.evidence_group / rec.filename
    wc.parent.mkdir(parents=True, exist_ok=True)
    wb = pyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "commodity", "lon", "lat"])
    ws.append(["OccA", "Au", 96.50, 45.50])
    ws.append(["OccB", "Cu", 96.52, 45.51])
    wb.save(wc)


def _inject_evidence(config):
    """Append a geology contact + a fault near the ingested #68 points so cells there score
    geology(20, contact-proximity) + structure(15) + geochem(15) = C-class prospects."""
    import geopandas as gpd
    from shapely.geometry import LineString, MultiLineString, MultiPolygon

    gpkg = _evidence_gpkg_03(config)
    pts = vector_io.read_layer(gpkg, "mineralized_points_point")
    reps = [g.representative_point() for g in pts.geometry]
    coords = [(p.x, p.y) for p in reps]
    epsg = config.target_epsg

    def _row(fid):
        r = dict.fromkeys(EVIDENCE_FIELDS, "")
        r["feature_id"] = fid
        r["validation_status"] = "Historical only"
        return r

    # geology polygon whose contact (boundary) sits within a grid cell of each point
    poly = pts.geometry.union_all().buffer(150)
    geol = poly if poly.geom_type == "MultiPolygon" else MultiPolygon([poly])
    gg = gpd.GeoDataFrame([_row("BUD-GEO50-0001")], geometry=[geol], crs=f"EPSG:{epsg}")
    vector_io.write_layer(
        gg.reindex(columns=[*EVIDENCE_FIELDS, "geometry"]),
        gpkg,
        layer="geology_units_50k_polygon",
        mode="a",
    )
    # a fault through the points
    line = (
        LineString(coords)
        if len(coords) >= 2
        else LineString([coords[0], (coords[0][0] + 300, coords[0][1])])
    )
    fg = gpd.GeoDataFrame(
        [_row("BUD-STR-0001")], geometry=[MultiLineString([line])], crs=f"EPSG:{epsg}"
    )
    vector_io.write_layer(
        fg.reindex(columns=[*EVIDENCE_FIELDS, "geometry"]),
        gpkg,
        layer="faults_structures_line",
        mode="a",
    )


# --------------------------------------------------------------------------- #
# unit: class bands + grid helpers
# --------------------------------------------------------------------------- #


def test_classify_bands():
    assert classify(75) == "A" and classify(74) == "B"
    assert classify(55) == "B" and classify(54) == "C"
    assert classify(35) == "C" and classify(34) == "D"
    assert classify(0) == "D"


def test_prospect_criteria_sum_100():
    assert sum(w for _k, w, _a in PROSPECT_CRITERIA) == 100


def test_make_grid_and_dissolve(project):
    import geopandas as gpd
    from shapely.geometry import box

    aoi = gpd.GeoDataFrame({"geometry": [box(0, 0, 1000, 1000)]}, crs="EPSG:32647")
    cells = vector_io.make_grid(aoi, 250, 32647)
    assert len(cells) == 16  # 4x4
    assert "grid_id" in cells.columns
    merged = vector_io.dissolve_adjacent(cells)
    assert len(merged) == 1  # all contiguous -> one cluster
    assert "cluster_id" in merged.columns


# --------------------------------------------------------------------------- #
# dry-run
# --------------------------------------------------------------------------- #


def test_phase04_scaffolds_five_folders(project):
    config, register, _tmp = project
    ctx = _ctx(config, register, dry_run=True)
    phase = Phase04ProspectRanking()
    phase.prepare(ctx)
    pdir = _p04_dir(config)
    subs = phase.custom_subfolders
    assert subs is not None and len(subs) == 5
    for sub in subs:
        assert (pdir / sub).is_dir(), f"missing folder {sub}"


def test_phase04_dry_run_templates(project):
    config, register, _tmp = project
    ctx = _ctx(config, register, dry_run=True)
    phase = Phase04ProspectRanking()
    phase.prepare(ctx)
    result = phase.run(ctx)
    assert result.status == "dry-run"
    pdir = _p04_dir(config)
    assert list((pdir / "03_Scoring_Matrix").glob("*Prospect_Ranking_Table*.xlsx"))
    assert list((pdir / "05_A_B_C_D_Field_Priority").glob("*Go_NoGo_Desktop_Decision_Matrix*.xlsx"))
    assert list((pdir / "04_Confidence_DataGap_NextAction").glob("*DataGap_and_NextAction*.xlsx"))
    assert list((pdir / "02_Prospect_Polygon_Delineation").glob("*Prospect_Polygons*.gpkg"))
    assert list((pdir / "05_A_B_C_D_Field_Priority").glob("*Method_Note.md"))


# --------------------------------------------------------------------------- #
# real run
# --------------------------------------------------------------------------- #


def test_phase04_real_run_grid_and_gate(raw_archive):
    # Full path on the bare synthetic archive: grid built, gate GO (provisional), data gaps recorded.
    config, register, _raw = raw_archive
    ctx = _ctx(config, register)
    _run_to_03(ctx)
    Phase03GeologySynthesis().run(ctx)

    phase = Phase04ProspectRanking()
    phase.prepare(ctx)
    result = phase.run(ctx)
    assert result.status == "ok"
    assert phase._grid_cells > 0  # a scoring grid was built over the AOI

    pdir = _p04_dir(config)
    assert list((pdir / "03_Scoring_Matrix").glob("*Evidence_Score_Grid*.gpkg"))

    # the three desktop-unavailable criteria are recorded as data gaps
    gap = next((pdir / "04_Confidence_DataGap_NextAction").glob("*DataGap_and_NextAction*.xlsx"))
    ws = openpyxl.load_workbook(gap).active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]
    gaps = {
        str(r[header.index("criterion")]) for r in data if "GAP" in str(r[header.index("status")])
    }
    assert {"rs", "field_pxrf", "drone"} <= gaps

    # gate advances but is provisional (human-completion items pending)
    report = phase.qaqc(ctx)
    decision = phase.gate(report, ctx)
    assert decision.status is GateStatus.GO
    assert decision.provisional


def test_phase04_delineates_prospects(raw_archive):
    # With ingested #68 points + a covering geology layer, cells score >= C and prospects form.
    config, register, _raw = raw_archive
    ctx = _ctx(config, register)
    _run_to_03(ctx)
    _write_68(ctx, config)
    Phase03GeologySynthesis().run(ctx)
    _inject_evidence(config)

    phase = Phase04ProspectRanking()
    phase.prepare(ctx)
    result = phase.run(ctx)
    assert result.status == "ok"
    assert phase._n_candidates >= 1

    pdir = _p04_dir(config)
    pp = next((pdir / "02_Prospect_Polygon_Delineation").glob("*Prospect_Polygons*.gpkg"))
    g = vector_io.read_layer(pp, "prospect_candidate_areas")
    assert len(g) >= 1
    assert set(g["prospect_class"]) <= {"A", "B", "C", "D"}
    # each polygon's class is consistent with its max_score band
    for _, r in g.iterrows():
        assert r["prospect_class"] == classify(r["max_score"])
    # candidate ids are the BUD-PSP scheme, ranked
    assert all(str(cid).startswith("BUD-PSP-") for cid in g["candidate_id"])

    # ranking table row count == polygon count
    rt = next((pdir / "03_Scoring_Matrix").glob("*Prospect_Ranking_Table*.xlsx"))
    ws = openpyxl.load_workbook(rt).active
    assert ws is not None
    assert ws.max_row == 1 + len(g)
